#!/usr/bin/env python3
"""
Standalone CLI to run LinkedIn and Reddit scrapers.
Outputs JSON to stdout or Excel file. Loads .env from current directory.

Usage:
  python main.py linkedin "keyword1 keyword2"
  python main.py reddit "keyword1 keyword2"
  python main.py both "keyword1 keyword2"
  python main.py linkedin "designer" --since 48
  python main.py reddit "web developer" --since 24
  python main.py both "designer" --format excel --output results.xlsx
"""
import argparse
import json
import os
import sys
from datetime import datetime

# Load .env before importing scrapers (they use os.getenv)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass


def _normalize_for_excel(posts: list) -> list:
    """Normalize posts from both platforms to a unified schema for Excel."""
    rows = []
    for p in posts:
        row = {
            "platform": p.get("platform", ""),
            "username": p.get("username", ""),
            "post_url": p.get("post_url", ""),
            "post_content": p.get("post_content", ""),
            "scraped_at": p.get("scraped_at", ""),
            "title": p.get("title", ""),
            "subreddit": p.get("subreddit", ""),
            "created_utc": p.get("created_utc", ""),
        }
        rows.append(row)
    return rows


def _export_excel(posts: list, output_path: str) -> None:
    """Export posts to an Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        print('{"posts": [], "error": "openpyxl not installed. Run: pip install openpyxl"}', file=sys.stderr)
        sys.exit(1)

    rows = _normalize_for_excel(posts)
    wb = Workbook()
    ws = wb.active
    ws.title = "Posts"

    headers = ["platform", "username", "post_url", "post_content", "scraped_at", "title", "subreddit", "created_utc"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            val = row_data.get(key, "")
            if val is None:
                val = ""
            ws.cell(row=row_idx, column=col_idx, value=str(val))

    # Auto-adjust column widths
    for col_idx, key in enumerate(headers, 1):
        max_len = len(key)
        for row_data in rows:
            val = row_data.get(key, "")
            max_len = min(max(max_len, len(str(val))), 80)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Run LinkedIn and/or Reddit scraper")
    parser.add_argument(
        "platform",
        choices=["linkedin", "reddit", "both"],
        help="Platform(s) to scrape: linkedin, reddit, or both",
    )
    parser.add_argument("keywords", type=str, help='Keywords (space-separated, e.g. "web developer designer")')
    parser.add_argument("--since", type=int, default=24, help="Time window in hours (default: 24)")
    parser.add_argument(
        "--format",
        choices=["json", "excel"],
        default="json",
        help="Output format: json (stdout) or excel file (default: json)",
    )
    parser.add_argument(
        "--output",
        "-o",
        type=str,
        default=None,
        help="Output file path (required for excel format; default: scraped_YYYYMMDD_HHMMSS.xlsx)",
    )
    args = parser.parse_args()

    keywords = [k.strip() for k in args.keywords.split() if k.strip()]
    if not keywords:
        print('{"posts": [], "error": "No keywords provided"}', file=sys.stderr)
        sys.exit(1)

    all_posts = []
    platforms_to_run = ["linkedin", "reddit"] if args.platform == "both" else [args.platform]

    try:
        for platform in platforms_to_run:
            if platform == "linkedin":
                from linkedin import scrape_linkedin
                posts = scrape_linkedin(keywords, quantity=50, since_hours=args.since)
            else:
                from reddit import scrape_reddit
                tf = "hour" if args.since <= 1 else "day" if args.since <= 24 else "week" if args.since <= 168 else "month" if args.since <= 720 else "year"
                posts = scrape_reddit(keywords, quantity=50, time_filter=tf)
            all_posts.extend(posts)

        if args.format == "excel":
            output_path = args.output
            if not output_path:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"scraped_{ts}.xlsx"
            _export_excel(all_posts, output_path)
            print(json.dumps({"posts": len(all_posts), "output_file": output_path}))
        else:
            output = {"posts": all_posts}
            print(json.dumps(output, indent=2))
    except Exception as e:
        print(json.dumps({"posts": [], "error": str(e)}), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
