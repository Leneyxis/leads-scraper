"""
Scraper API — JSON endpoints for LinkedIn and Reddit.
Run: uvicorn api:app --reload
Docs: http://localhost:8000/docs
"""
import os
from datetime import datetime
from typing import Any, List, Optional

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field
from io import BytesIO

# Load .env before importing scrapers
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# Configure logging before scrapers (so [linkedin] / [reddit] show correctly)
import logging
logging.basicConfig(level=logging.INFO, format="[%(name)s] %(message)s")

from reddit import scrape_reddit
from linkedin import scrape_linkedin

app = FastAPI(
    title="LinkedIn & Reddit Scraper API",
    description="Scrape hiring posts from LinkedIn and Reddit. Returns JSON.",
    version="1.0.0",
)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])


# ---------------------------------------------------------------------------
# Response models (consistent JSON schema)
# ---------------------------------------------------------------------------

class Post(BaseModel):
    """Unified post schema for both platforms."""
    platform: str = Field(..., description="linkedin | reddit")
    username: str = Field(..., description="Author username")
    post_url: str = Field(..., description="URL to the post")
    post_content: str = Field(..., description="Post body/text")
    scraped_at: str = Field(..., description="ISO timestamp")
    title: Optional[str] = Field(None, description="Post title (Reddit)")
    subreddit: Optional[str] = Field(None, description="Subreddit (Reddit)")
    created_utc: Optional[Any] = Field(None, description="Reddit created_utc")


class ScrapeResponse(BaseModel):
    """Standard JSON response for scrape endpoints."""
    success: bool = Field(..., description="Whether the scrape completed without error")
    platform: str = Field(..., description="linkedin | reddit | both")
    count: int = Field(..., description="Number of posts returned")
    posts: List[Post] = Field(default_factory=list, description="List of posts")
    error: Optional[str] = Field(None, description="Error message if success=False")


class ScrapeRequest(BaseModel):
    """Request body for POST /api/scrape."""
    platform: str = Field("both", description="reddit | linkedin | both")
    keywords: List[str] = Field(..., min_length=1, description="Search keywords")
    quantity: int = Field(50, ge=1, le=200, description="Max posts to fetch")
    time_filter: str = Field("day", description="Reddit: hour, day, week, month, year, all")
    since_hours: int = Field(24, description="LinkedIn time window (24, 168, 720)")


class ExportExcelRequest(BaseModel):
    posts: List[dict]


def _to_post(p: dict) -> Post:
    """Normalize scraper output to Post model."""
    return Post(
        platform=p.get("platform", ""),
        username=p.get("username", ""),
        post_url=p.get("post_url", ""),
        post_content=p.get("post_content", ""),
        scraped_at=p.get("scraped_at", ""),
        title=p.get("title"),
        subreddit=p.get("subreddit"),
        created_utc=p.get("created_utc"),
    )


def _normalize_for_excel(posts: list) -> list:
    rows = []
    for p in posts:
        rows.append({
            "platform": p.get("platform", ""),
            "username": p.get("username", ""),
            "post_url": p.get("post_url", ""),
            "post_content": p.get("post_content", ""),
            "scraped_at": p.get("scraped_at", ""),
            "title": p.get("title", ""),
            "subreddit": p.get("subreddit", ""),
            "created_utc": p.get("created_utc", ""),
        })
    return rows


def _build_excel_bytes(posts: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    rows = _normalize_for_excel(posts)
    wb = Workbook()
    ws = wb.active
    ws.title = "Posts"
    headers = ["platform", "username", "post_url", "post_content", "scraped_at", "title", "subreddit", "created_utc"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            val = row_data.get(key, "") or ""
            ws.cell(row=row_idx, column=col_idx, value=str(val))
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _linkedin_configured() -> bool:
    return bool(
        os.getenv("LINKEDIN_LI_AT", "").strip()
        or os.getenv("LINKEDIN_COOKIES", "").strip()
        or (os.getenv("LINKEDIN_EMAIL", "").strip() and os.getenv("LINKEDIN_PASSWORD", "").strip())
    )


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@app.get("/api")
def api_info():
    """API info and available endpoints."""
    return {
        "name": "LinkedIn & Reddit Scraper API",
        "version": "1.0.0",
        "docs": "/docs",
        "endpoints": {
            "GET /api/linkedin": "Scrape LinkedIn (query: keywords, quantity)",
            "GET /api/reddit": "Scrape Reddit (query: keywords, quantity, time_filter)",
            "POST /api/scrape": "Scrape one or both (body: platform, keywords, quantity)",
            "GET /api/health": "Health check",
        },
    }


@app.get("/api/linkedin", response_model=ScrapeResponse)
def get_linkedin(
    keywords: str = Query(..., description="Search keywords (space-separated)"),
    quantity: int = Query(50, ge=1, le=200, description="Max posts"),
) -> ScrapeResponse:
    """Scrape LinkedIn posts. Returns JSON."""
    kw_list = [k.strip() for k in keywords.split() if k.strip()]
    if not kw_list:
        raise HTTPException(status_code=400, detail="keywords is required")
    try:
        posts = scrape_linkedin(keywords=kw_list, quantity=quantity)
        return ScrapeResponse(
            success=True,
            platform="linkedin",
            count=len(posts),
            posts=[_to_post(p) for p in posts],
        )
    except Exception as e:
        return ScrapeResponse(
            success=False,
            platform="linkedin",
            count=0,
            posts=[],
            error=str(e),
        )


@app.get("/api/reddit", response_model=ScrapeResponse)
def get_reddit(
    keywords: str = Query(..., description="Search keywords (space-separated)"),
    quantity: int = Query(50, ge=1, le=200, description="Max posts"),
    time_filter: str = Query("day", description="hour, day, week, month, year, all"),
) -> ScrapeResponse:
    """Scrape Reddit [Hiring] posts. Returns JSON."""
    kw_list = [k.strip() for k in keywords.split() if k.strip()]
    if not kw_list:
        raise HTTPException(status_code=400, detail="keywords is required")
    try:
        posts = scrape_reddit(keywords=kw_list, quantity=quantity, time_filter=time_filter)
        return ScrapeResponse(
            success=True,
            platform="reddit",
            count=len(posts),
            posts=[_to_post(p) for p in posts],
        )
    except Exception as e:
        return ScrapeResponse(
            success=False,
            platform="reddit",
            count=0,
            posts=[],
            error=str(e),
        )


@app.post("/api/scrape", response_model=ScrapeResponse)
def post_scrape(req: ScrapeRequest) -> ScrapeResponse:
    """Scrape LinkedIn and/or Reddit. Returns JSON."""
    platforms = ["linkedin", "reddit"] if req.platform == "both" else [req.platform]
    all_posts = []
    try:
        for p in platforms:
            if p == "reddit":
                posts = scrape_reddit(
                    keywords=req.keywords,
                    quantity=req.quantity,
                    time_filter=req.time_filter,
                )
            else:
                posts = scrape_linkedin(
                    keywords=req.keywords,
                    quantity=req.quantity,
                    since_hours=req.since_hours,
                )
            all_posts.extend(posts)
        return ScrapeResponse(
            success=True,
            platform=req.platform,
            count=len(all_posts),
            posts=[_to_post(p) for p in all_posts],
        )
    except Exception as e:
        return ScrapeResponse(
            success=False,
            platform=req.platform,
            count=0,
            posts=[],
            error=str(e),
        )


@app.post("/api/export-excel")
def export_excel(req: ExportExcelRequest):
    """Export posts to Excel file."""
    try:
        data = _build_excel_bytes(req.posts or [])
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"scraped_{ts}.xlsx"
        return StreamingResponse(
            BytesIO(data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")


@app.get("/api/health")
def health():
    """Health check and configuration status."""
    return {
        "status": "ok",
        "reddit_configured": bool(os.getenv("REDDIT_CLIENT_ID", "").strip()),
        "linkedin_configured": _linkedin_configured(),
    }


# Serve frontend (do NOT mount StaticFiles at "/" — POST /api/* can hit that mount and
# StaticFiles only allows GET/HEAD → 405 Method Not Allowed behind some proxies/tunnels).
frontend_path = os.path.join(os.path.dirname(__file__), "frontend")
if os.path.exists(frontend_path):
    _index_html = os.path.join(frontend_path, "index.html")

    @app.get("/")
    def serve_index():
        return FileResponse(_index_html)
